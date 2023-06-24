import openpyxl
from Utilities.Read_Config_File import Read_Confi_File


class Read_Excel_File:

    @staticmethod
    def read_excel(rownum, colnum):
        R = Read_Confi_File()
        Excel = openpyxl.load_workbook(R.get_Excel_Path())
        Sheet = Excel.active
        return Sheet.cell(row=rownum, column=colnum).value

    @staticmethod
    def write_excel(rownum, colnum, data):
        R = Read_Confi_File()
        Book = openpyxl.load_workbook(R.get_Excel_Path())
        Sheet = Book.active
        Sheet.cell(row=rownum, column=colnum).value = data
        Book.save(R.get_Excel_Path())